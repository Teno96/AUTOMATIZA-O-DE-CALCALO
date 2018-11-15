'''
SE TODAS AS INFORMAÇÕES ESTIVEREM NA PLANILHA DO EXCEL, O OBJETIVO É COLAR AS INFORMAÇÕES QUE ESTÃO NA PLANILHA PARA A PLANILHA DE CÁLCULOS E TIRAR PRINT.
'''

'ABRIR PLANILHA COM A LISTA DOS POSTES'

''' ISSO TUDO SENDO UM LOOP
'COMANDO PARA COPIAR E COLAR CADA INFORMAÇÃO DA LISTA PARA A PLANILHA DE CÁLCULO'

'TIRAR PRINT DA TELA'

'ABRIR PAINT'

'COLAR A IMAGEM NO PAINT E INSERIR O NUMERO DO POSTE QUANDO SALVAR'

'FECHAR PAINT' '''

import os
import pyautogui
import pandas as pd
import xlrd
import xlwt
import openpyxl

postes=pd.read_excel('C:\\Users\\heito\\Documents\\postes.xlsx')
book=xlrd.open_workbook('C:\\Users\\heito\\Documents\\postes.xlsx')
book_a=openpyxl.load_workbook("C:\\Users\\heito\\Documents\\Programa de Cálculo de Traçãooo.xlsm")
'book_a=xlwt.Workbook("C:\\Users\\heito\\Documents\\Programa de Cálculo de Traçãooo.xls")'
sheet=book.sheet_by_index(0)

'sheet_a=book_a.active'
sheet_a=book_a.get_sheet_by_name('Dados de Entrada')
'cell=int(sheet.cell(1,0).value)'

i=int(postes['numero postes'].max())

for x in range(i):
    x=x+1
    z=sheet.cell((x),3).value
    if z != 0:
         h=int(sheet.cell((x),2).value)
         vao_a=int(sheet.cell((x),4).value)
         vao_b=int(sheet.cell((x),5).value)
         fase_mt_a=int(sheet.cell((x),6).value)
         mt_a=sheet.cell((x),7).value
         mt_b=sheet.cell((x),8).value
         fase_mt_b=int(sheet.cell((x),9).value)
         fase_bt_a=int(sheet.cell((x),10).value)
         neutro_a=sheet.cell((x),11).value
         bt_a=sheet.cell((x),12).value
         bt_b=sheet.cell((x),13).value
         neutro_b=sheet.cell((x),14).value
         fase_bt_b=int(sheet.cell((x),15).value)
         CA=sheet.cell((x),16).value
         CA_b=sheet.cell((x),17).value
         oc_a=sheet.cell((x),18).value
         oc_b=sheet.cell((x),19).value

         os.startfile("C:\\Users\\heito\\Documents\\Programa de Cálculo de Traçãooo.xls")
         
               
         
         sheet_a['F16']=vao_a
         sheet_a['K16']=vao_b
         'xlrd.formula.cellname(9,5)'
         
         '''sheet_a.write(11,11,z)
         sheet_a.write(15,5,vao_a)
         sheet_a.write(15,10,vao_b)
         sheet_a.write(19,1,fase_mt_a)
         sheet_a.write(19,5,mt_a)
         sheet_a.write(19,10,mt_b)
         sheet_a.write(19,14,fase_mt_b)
         sheet_a.write(24,1,fase_bt_a)
         sheet_a.write(24,3,neutro_a)
         sheet_a.write(24,5,bt_a)
         sheet_a.write(24,10,bt_b)
         sheet_a.write(24,12,neutro_b)
         sheet_a.write(24,14,fase_bt_b)
         sheet_a.write(26,5,CA)
         sheet_a.write(26,10,CA_b)
         sheet_a.write(32,5,oc_a)
         sheet_a.write(32,10,oc_b)'''


         '''pyautogui.hotkey('win', 'ç')
         pyautogui.click(x=435, y=191, interval=1)
         pyautogui.doubleClick(x=435, y=190, interval=1)
         pyautogui.typewrite(str(h))
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(z))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(vao_a))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(vao_b))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(fase_mt_a))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(mt_a))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(mt_b))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(fase_mt_b))
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(fase_bt_a))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(neutro_a))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(bt_a))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(bt_b))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(neutro_b))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(fase_bt_b))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(CA))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(CA_b))
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(oc_a))
         pyautogui.hotkey('tab')
         pyautogui.typewrite(str(oc_b))'''


         pyautogui.screenshot('C:\\Users\\heito\\Work\\Fotos\\POSTE '+ str(i) +'.jpeg')

    else:
        next








'''book_a=openpyxl.Workbook("C:\\Users\\heito\\Documents\\Programa de Cálculo de Tração.xlsx")
         sheet_a=book_a.active
         sheet_a['F16']=vao_a
         xlrd.formula.cellname(9,5)
         sheet_a.write(9,5,h)
         sheet_a.write(11,11,z)
         sheet_a.write(15,5,vao_a)
         sheet_a.write(15,10,vao_b)
         sheet_a.write(19,1,fase_mt_a)
         sheet_a.write(19,5,mt_a)
         sheet_a.write(19,10,mt_b)
         sheet_a.write(19,14,fase_mt_b)
         sheet_a.write(24,1,fase_bt_a)
         sheet_a.write(24,3,neutro_a)
         sheet_a.write(24,5,bt_a)
         sheet_a.write(24,10,bt_b)
         sheet_a.write(24,12,neutro_b)
         sheet_a.write(24,14,fase_bt_b)
         sheet_a.write(26,5,CA)
         sheet_a.write(26,10,CA_b)
         sheet_a.write(32,5,oc_a)
         sheet_a.write(32,10,oc_b)'''

